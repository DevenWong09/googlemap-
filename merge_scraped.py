#!/usr/bin/env python3
"""
读取「google map 地址抓取」文件夹内全部 JSON，去重后自动汇总更新「地址表.xlsx」。

默认直接运行即可，无需额外参数：
    python merge_scraped.py
"""

from __future__ import annotations

import argparse
import json
import re
import sys
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font

SCRIPT_DIR = Path(__file__).resolve().parent
PROJECT_DIR = SCRIPT_DIR.parent
DEFAULT_DATA_DIR = PROJECT_DIR / "google map 地址抓取"
DEFAULT_EXCEL = PROJECT_DIR / "地址表.xlsx"

JSON_COLUMNS = [
    "name",
    "address",
    "phone",
    "rating",
    "website",
    "maps_url",
    "search_query",
    "page_title",
    "page_host",
    "scraped_at",
    "source_file",
]

EXTRA_COLUMNS = ["匹配状态"]


def normalize(text: str) -> str:
    return re.sub(r"\s+", "", str(text or "").lower())


def dedupe_key(item: dict) -> str:
    name = normalize(item.get("name", ""))
    address = normalize(item.get("address", ""))
    if name or address:
        return f"{name}|{address}"
    return normalize(item.get("source_file", ""))


def parse_scraped_at(item: dict) -> datetime:
    raw = str(item.get("scraped_at", "")).strip()
    if not raw:
        return datetime.min.replace(tzinfo=timezone.utc)
    try:
        return datetime.fromisoformat(raw.replace("Z", "+00:00"))
    except ValueError:
        return datetime.min.replace(tzinfo=timezone.utc)


def dedupe_json_records(records: list[dict]) -> list[dict]:
    """同一门店（name+address 相同）只保留 scraped_at 最新的一条。"""
    best: dict[str, dict] = {}
    for item in records:
        key = dedupe_key(item)
        existing = best.get(key)
        if existing is None or parse_scraped_at(item) > parse_scraped_at(existing):
            best[key] = item

    deduped = list(best.values())
    removed = len(records) - len(deduped)
    if removed:
        print(f"JSON 去重：{len(records)} → {len(deduped)}（移除 {removed} 条重复）")
    return deduped


def load_json_files(data_dir: Path) -> list[dict]:
    if not data_dir.is_dir():
        raise FileNotFoundError(f"未找到数据文件夹：{data_dir}")

    records: list[dict] = []
    for path in sorted(data_dir.glob("*.json")):
        if path.name.startswith("maps-debug") or "诊断" in path.name:
            print(f"跳过诊断文件：{path.name}")
            continue

        try:
            with path.open(encoding="utf-8") as f:
                data = json.load(f)
        except (json.JSONDecodeError, OSError) as exc:
            print(f"跳过无效文件 {path.name}：{exc}")
            continue

        if not isinstance(data, dict):
            print(f"跳过非对象 JSON：{path.name}")
            continue

        if "extracted" in data and "name" not in data:
            print(f"跳过诊断文件：{path.name}")
            continue

        if not data.get("name"):
            print(f"跳过无 name 字段的 JSON：{path.name}")
            continue

        data["source_file"] = path.name
        records.append(data)

    return dedupe_json_records(records)


def match_score(store_name: str, item: dict) -> int:
    target = normalize(store_name)
    if not target:
        return 0

    candidates = [
        (item.get("name", ""), 100),
        (item.get("search_query", ""), 80),
        (item.get("source_file", "").rsplit(".", 1)[0], 60),
    ]
    best = 0
    for candidate, base in candidates:
        norm = normalize(candidate)
        if not norm:
            continue
        if norm == target:
            best = max(best, base + 50)
        elif target in norm or norm in target:
            best = max(best, base)
    return best


def json_to_cells(item: dict) -> dict[str, object]:
    return {col: item.get(col, "") for col in JSON_COLUMNS}


def write_json_fields(ws, row_num: int, col_map: dict[str, int], item: dict) -> None:
    for col_name, value in json_to_cells(item).items():
        col = col_map.get(col_name)
        if col:
            ws.cell(row=row_num, column=col, value=value)


def get_cell_str(ws, row_num: int, col: int | None) -> str:
    if not col:
        return ""
    value = ws.cell(row=row_num, column=col).value
    return str(value).strip() if value is not None else ""


def row_dedupe_key_from_sheet(
    ws,
    row_num: int,
    col_map: dict[str, int],
    legacy_name_col: int | None,
    legacy_addr_col: int | None,
) -> str:
    name_col = col_map.get("name") or legacy_name_col
    addr_col = col_map.get("address") or legacy_addr_col
    return dedupe_key({
        "name": get_cell_str(ws, row_num, name_col),
        "address": get_cell_str(ws, row_num, addr_col),
    })


def find_row_by_source_file(ws, col_map: dict[str, int], source_file: str) -> int | None:
    col = col_map.get("source_file")
    if not col or not source_file:
        return None
    for row_num in range(2, ws.max_row + 1):
        if get_cell_str(ws, row_num, col) == source_file:
            return row_num
    return None


def find_row_by_dedupe_key(
    ws,
    col_map: dict[str, int],
    item: dict,
    legacy_name_col: int | None,
    legacy_addr_col: int | None,
) -> int | None:
    target = dedupe_key(item)
    if not target or target == "|":
        return None
    for row_num in range(2, ws.max_row + 1):
        if row_dedupe_key_from_sheet(ws, row_num, col_map, legacy_name_col, legacy_addr_col) == target:
            return row_num
    return None


def fill_row(
    ws,
    row_num: int,
    col_map: dict[str, int],
    item: dict,
    status: str,
    legacy_name_col: int | None,
    legacy_addr_col: int | None,
) -> None:
    write_json_fields(ws, row_num, col_map, item)
    ws.cell(row=row_num, column=col_map["匹配状态"], value=status)
    if legacy_name_col:
        ws.cell(row=row_num, column=legacy_name_col, value=item.get("name", ""))
    if legacy_addr_col:
        ws.cell(row=row_num, column=legacy_addr_col, value=item.get("address", ""))


def merge_with_excel(input_path: Path, scraped: list[dict], output_path: Path) -> None:
    wb = load_workbook(input_path)
    ws = wb.active
    header = [str(c).strip() if c is not None else "" for c in next(ws.iter_rows(max_row=1, values_only=True))]

    legacy_name_idx = next((i for i, h in enumerate(header) if h == "店名"), None)
    if legacy_name_idx is None and "name" not in header:
        raise ValueError("Excel 必须包含「店名」或「name」列")

    legacy_addr_idx = next((i for i, h in enumerate(header) if h == "地址"), None)
    seq_idx = next((i for i, h in enumerate(header) if h in ("序号", "seq")), None)

    legacy_name_col = legacy_name_idx + 1 if legacy_name_idx is not None else None
    legacy_addr_col = legacy_addr_idx + 1 if legacy_addr_idx is not None else None

    out_header = header[:]
    for col in JSON_COLUMNS + EXTRA_COLUMNS:
        if col not in out_header:
            out_header.append(col)

    for col_idx, title in enumerate(out_header, start=1):
        cell = ws.cell(row=1, column=col_idx, value=title)
        cell.font = Font(bold=True)

    col_map = {title: idx + 1 for idx, title in enumerate(out_header)}

    updated = 0
    matched = 0
    appended = 0
    skipped = 0
    used_rows: set[int] = set()

    def get_store_name(row_num: int) -> str:
        if "name" in col_map:
            name = get_cell_str(ws, row_num, col_map["name"])
            if name:
                return name
        if legacy_name_col:
            return get_cell_str(ws, row_num, legacy_name_col)
        return ""

    # ① 按 source_file / name+address 更新已有行（重复运行不追加）
    for item in scraped:
        source = item.get("source_file", "")
        row_num = find_row_by_source_file(ws, col_map, source)
        if row_num is None:
            row_num = find_row_by_dedupe_key(ws, col_map, item, legacy_name_col, legacy_addr_col)
        if row_num is not None:
            fill_row(ws, row_num, col_map, item, "已更新", legacy_name_col, legacy_addr_col)
            used_rows.add(row_num)
            updated += 1

    # ② 按 Excel「店名」列匹配尚未写入的 JSON
    for row_num in range(2, ws.max_row + 1):
        if row_num in used_rows:
            continue
        store_name = get_store_name(row_num)
        if not store_name:
            continue

        best_item = None
        best_score = 0
        for item in scraped:
            source = item.get("source_file", "")
            if find_row_by_source_file(ws, col_map, source):
                continue
            if find_row_by_dedupe_key(ws, col_map, item, legacy_name_col, legacy_addr_col):
                continue
            score = match_score(store_name, item)
            if score > best_score:
                best_score = score
                best_item = item

        if best_item:
            fill_row(ws, row_num, col_map, best_item, "已匹配", legacy_name_col, legacy_addr_col)
            used_rows.add(row_num)
            matched += 1
        elif get_cell_str(ws, row_num, col_map.get("匹配状态")) == "":
            ws.cell(row=row_num, column=col_map["匹配状态"], value="未匹配")

    # ③ 追加 Excel 中尚不存在的新 JSON
    next_seq = 1
    if seq_idx is not None:
        seq_col = seq_idx + 1
        for row_num in range(2, ws.max_row + 1):
            value = ws.cell(row=row_num, column=seq_col).value
            if isinstance(value, (int, float)):
                next_seq = max(next_seq, int(value) + 1)

    for item in scraped:
        source = item.get("source_file", "")
        if find_row_by_source_file(ws, col_map, source):
            continue
        if find_row_by_dedupe_key(ws, col_map, item, legacy_name_col, legacy_addr_col):
            skipped += 1
            continue

        row_num = ws.max_row + 1
        if seq_idx is not None:
            ws.cell(row=row_num, column=seq_idx + 1, value=next_seq)
            next_seq += 1
        fill_row(ws, row_num, col_map, item, "JSON追加", legacy_name_col, legacy_addr_col)
        appended += 1

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    wb.close()

    total_rows = ws.max_row - 1
    print(
        f"Excel 更新完成：{updated} 条已更新，{matched} 条新匹配，"
        f"{appended} 条追加，{skipped} 条重复跳过，共 {total_rows} 行"
    )
    print(f"结果已保存：{output_path}")


def export_all_scraped(scraped: list[dict], output_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "抓取结果"

    headers = JSON_COLUMNS + ["匹配状态"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for item in scraped:
        row = [item.get(col, "") for col in JSON_COLUMNS]
        row.append("已抓取")
        ws.append(row)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    print(f"已汇总 {len(scraped)} 个 JSON → {output_path}")


def main() -> None:
    parser = argparse.ArgumentParser(
        description="读取 JSON 文件夹，去重后自动汇总更新 Excel（默认更新 地址表.xlsx）",
    )
    parser.add_argument(
        "-d", "--data-dir",
        default=str(DEFAULT_DATA_DIR),
        help=f"JSON 文件夹（默认：{DEFAULT_DATA_DIR}）",
    )
    parser.add_argument(
        "-i", "--input",
        default=str(DEFAULT_EXCEL),
        help=f"Excel 路径（默认：{DEFAULT_EXCEL}）",
    )
    parser.add_argument(
        "-o", "--output",
        default="",
        help="输出 Excel 路径（默认与 -i 相同，直接更新原表）",
    )
    args = parser.parse_args()

    data_dir = Path(args.data_dir)
    scraped = load_json_files(data_dir)

    if not scraped:
        print(f"文件夹中没有有效 JSON：{data_dir}")
        sys.exit(1)

    print(f"有效 JSON：{len(scraped)} 条")

    input_path = Path(args.input)
    output_path = Path(args.output) if args.output else input_path

    if input_path.exists():
        merge_with_excel(input_path, scraped, output_path)
    else:
        print(f"未找到 {input_path}，将仅根据 JSON 创建新表")
        export_all_scraped(scraped, output_path)


if __name__ == "__main__":
    main()
